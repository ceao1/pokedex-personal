from contextlib import contextmanager
from pathlib import Path

import httpx
import pytest
from openpyxl import Workbook

from pokedex.catalog.models import Card, CardRef
from pokedex.catalog.service import CatalogService
from pokedex.wishlist import repository
from pokedex.wishlist.excel import SHEET_DEX, SHEET_GALLERY
from pokedex.wishlist.service import ImportService

XLSX = Path(__file__).parents[3] / "Pokedex_Viviente_151.xlsx"


class FakeCatalogService:
    """Espeja cartas inventadas sin tocar la red ni TCGdex."""

    def __init__(self, conn):
        self._conn = conn
        self.mirrored = []

    async def get_card(self, card_id: str):
        return None

    async def find_by_set_and_number(self, set_id: str, local_id: str):
        return None

    async def list_set_cards(self, set_id: str):
        return []


class FakeCatalogPortConNumeros:
    """Puerto falso que sí resuelve números fijos del set 151, para probar
    que la galería encuentra y marca la misma carta que ya resolvió una
    opción, en vez de crear una fila nueva."""

    def __init__(self, cartas: dict[str, str]):
        self._cartas = cartas  # local_id -> card_id

    async def get_card(self, card_id: str):
        return None

    async def find_by_set_and_number(self, set_id: str, local_id: str):
        card_id = self._cartas.get(local_id) if set_id == "sv03.5" else None
        if card_id is None:
            return None
        return Card(
            id=card_id, name="Ditto", set_id=set_id, set_name="151", local_id=local_id, raw={}
        )

    async def list_set_cards(self, set_id: str):
        return []


class FakeCatalogPortVintage:
    """Puerto falso que resuelve un Pokémon vintage vía `list_set_cards`
    (opción 3), sin tocar la red. A diferencia de `FakeCatalogService`, sí
    implementa `get_card` -- lo necesita el fix de este round, que debe
    espejar la carta antes de insertarla como wishlist_item."""

    def __init__(self):
        self._card = Card(
            id="base1-4",
            name="Ditto",
            set_id="base1",
            set_name="Base Set",
            local_id="4",
            raw={},
        )

    async def get_card(self, card_id: str):
        return self._card if card_id == "base1-4" else None

    async def find_by_set_and_number(self, set_id: str, local_id: str):
        return None

    async def list_set_cards(self, set_id: str):
        if set_id != "base1":
            return []
        return [CardRef(id="base1-4", local_id="4", name="Ditto")]


def _build_vintage_workbook(tmp_path: Path) -> Path:
    """Un Pokémon cuya única opción es la 3 (vintage), resuelta por
    `list_set_cards` -- el camino que no espeja nada por sí solo."""
    workbook = Workbook()
    dex_sheet = workbook.active
    dex_sheet.title = SHEET_DEX
    dex_sheet["A4"] = 132
    dex_sheet["B4"] = "Ditto"
    dex_sheet["M4"] = "Ditto Base Set"
    dex_sheet["N4"] = "5.00"
    workbook.create_sheet(SHEET_GALLERY)

    path = tmp_path / "vintage.xlsx"
    workbook.save(path)
    return path


class FakeCatalogPortSiempreFalla:
    """Simula TCGdex completamente caído: toda consulta explota con un
    error de red, nunca con un 404/None."""

    async def get_card(self, card_id: str):
        raise httpx.ConnectTimeout("catálogo caído (fake)")

    async def find_by_set_and_number(self, set_id: str, local_id: str):
        raise httpx.ConnectTimeout("catálogo caído (fake)")

    async def list_set_cards(self, set_id: str):
        return []


class FakeCatalogPortConFallas:
    """Puerto falso configurable: resuelve los `local_id` en `cartas` a una
    carta, hace explotar los que están en `fallan` (simulando TCGdex caído
    para esa consulta puntual) y devuelve `None` real para el resto (un
    "no existe" genuino, no un error)."""

    def __init__(self, cartas: dict[str, str] | None = None, fallan: set[str] | None = None):
        self._cartas = cartas or {}  # local_id -> card_id
        self._fallan = fallan or set()

    async def get_card(self, card_id: str):
        for local_id, cid in self._cartas.items():
            if cid == card_id:
                return Card(
                    id=cid, name="Ditto", set_id="sv03.5", set_name="151", local_id=local_id, raw={}
                )
        return None

    async def find_by_set_and_number(self, set_id: str, local_id: str):
        if local_id in self._fallan:
            raise httpx.ConnectTimeout("catálogo caído (fake)")
        card_id = self._cartas.get(local_id) if set_id == "sv03.5" else None
        if card_id is None:
            return None
        return Card(
            id=card_id, name="Ditto", set_id=set_id, set_name="151", local_id=local_id, raw={}
        )

    async def list_set_cards(self, set_id: str):
        return []


def _build_dos_pokemon_opcion1(tmp_path: Path) -> Path:
    """Dos filas con solo opción 1 (sin opción 2/3/4 ni galería), para
    aislar el comportamiento de catálogo inalcanzable del resto del
    pipeline."""
    workbook = Workbook()
    dex_sheet = workbook.active
    dex_sheet.title = SHEET_DEX
    dex_sheet["A4"] = 1
    dex_sheet["B4"] = "Ditto"
    dex_sheet["E4"] = "Ditto 001/165"
    dex_sheet["G4"] = "0.10"
    dex_sheet["A5"] = 2
    dex_sheet["B5"] = "Clon"
    dex_sheet["E5"] = "Clon 002/165"
    dex_sheet["G5"] = "0.20"
    workbook.create_sheet(SHEET_GALLERY)

    path = tmp_path / "dos_pokemon.xlsx"
    workbook.save(path)
    return path


def _build_mini_workbook(tmp_path: Path) -> Path:
    """Un Pokémon con opción 1 (001/165) y opción 2 (166/165, no-reverse:
    una Illustration Rare propia), más una fila de galería que la nombra por
    el mismo número — el caso real es "Bulbasaur 151 166/165"."""
    workbook = Workbook()
    dex_sheet = workbook.active
    dex_sheet.title = SHEET_DEX
    dex_sheet["A4"] = 1
    dex_sheet["B4"] = "Ditto"
    dex_sheet["E4"] = "Ditto 001/165"
    dex_sheet["G4"] = "0.10"
    dex_sheet["I4"] = "Ditto 166/165"
    dex_sheet["K4"] = "12.00"

    gallery_sheet = workbook.create_sheet(SHEET_GALLERY)
    gallery_sheet["A4"] = 1
    gallery_sheet["B4"] = "Ditto"
    gallery_sheet["C4"] = "Ditto 151 166/165"
    gallery_sheet["D4"] = "12.00"

    path = tmp_path / "mini.xlsx"
    workbook.save(path)
    return path


@pytest.fixture()
def conn_factory(clean_db):
    @contextmanager
    def factory():
        yield clean_db

    return factory


async def test_el_import_siembra_los_151_pokemon(conn_factory, clean_db):
    service = ImportService(FakeCatalogService(clean_db), conn_factory)
    resumen = await service.import_workbook(XLSX)
    assert resumen.pokemon == 151
    filas = repository.list_pokedex(clean_db)
    assert len(filas) == 151
    assert filas[0]["name"] == "Bulbasaur"
    assert filas[-1]["name"] == "Mew"


async def test_el_import_no_crea_ejemplares(conn_factory, clean_db):
    """Restricción global: la columna ✔ del Excel se ignora por completo.

    `app.owned_copy` no existe en ningún plan ejecutado todavía (llega con el
    plan de captura). Este test se deja en skip, con la tabla nombrada
    explícitamente, para que se vuelva un test real y ejecutable el día que
    esa tabla exista — mientras tanto la restricción sigue cubierta por
    `test_la_columna_de_check_se_ignora` en `tests/wishlist/test_excel.py`.
    """
    pytest.skip("app.owned_copy no existe todavía (plan de captura pendiente)")
    service = ImportService(FakeCatalogService(clean_db), conn_factory)
    await service.import_workbook(XLSX)
    total = clean_db.execute("select count(*) as n from app.owned_copy").fetchone()["n"]
    assert total == 0


async def test_reimportar_es_idempotente(conn_factory, clean_db):
    service = ImportService(FakeCatalogService(clean_db), conn_factory)
    primero = await service.import_workbook(XLSX)
    despues_del_primero = len(repository.list_wishlist(clean_db))
    segundo = await service.import_workbook(XLSX)
    assert len(repository.list_wishlist(clean_db)) == despues_del_primero
    assert segundo.items_creados == 0
    assert primero.items_creados > 0


async def test_una_carta_vintage_se_espeja_antes_del_upsert(conn_factory, clean_db, tmp_path):
    """Bug real, encontrado corriendo el import de verdad: la opción 3
    (vintage) resuelve por `list_set_cards`, que devuelve un `CardRef`
    liviano (id, localId, name) y nunca espeja nada en `app.card` -- a
    diferencia de las opciones 1 y 2, que resuelven por
    `find_by_set_and_number` y `CatalogService` las espeja como efecto
    secundario. Sin espejar la carta antes del upsert, el FK de
    `wishlist_item.card_id` la rechaza con `ForeignKeyViolation` la primera
    vez que aparece una carta vintage nueva -- no seedeada aquí a propósito,
    para reproducir el fallo real."""
    path = _build_vintage_workbook(tmp_path)
    catalog = CatalogService(FakeCatalogPortVintage(), conn_factory)
    service = ImportService(catalog, conn_factory)

    resumen = await service.import_workbook(path)

    assert resumen.sin_resolver == 0
    filas = repository.list_wishlist(clean_db, dex_number=132)
    assert filas[0]["card_id"] == "base1-4"
    espejada = clean_db.execute(
        "select count(*) as n from app.card where id = 'base1-4'"
    ).fetchone()["n"]
    assert espejada == 1


async def test_la_galeria_marca_favoritos_en_vez_de_duplicar(conn_factory, clean_db, tmp_path):
    """Cuando la galería resuelve a la misma carta que ya insertó una opción,
    no debe crear una fila nueva: debe marcar `is_favorite` en la existente."""
    path = _build_mini_workbook(tmp_path)
    port = FakeCatalogPortConNumeros({"001": "sv03.5-001", "166": "sv03.5-166"})
    catalog = CatalogService(port, conn_factory)
    service = ImportService(catalog, conn_factory)

    await service.import_workbook(path)

    filas = repository.list_wishlist(clean_db, dex_number=1)
    assert len(filas) == 2, "la galería no debe agregar una fila además de opción 1 y 2"
    opcion_2 = next(f for f in filas if f["card_id"] == "sv03.5-166")
    assert opcion_2["is_favorite"] is True


async def test_la_galeria_marca_favorita_una_fila_sin_resolver_cuando_el_texto_no_resuelve(
    conn_factory, clean_db
):
    """Trece filas de la galería dicen literalmente 'Ya está en tu Opción 2':
    no traen número, así que no hay con qué fusionar. Deben seguir quedando
    como una fila propia, marcada como favorita."""
    service = ImportService(FakeCatalogService(clean_db), conn_factory)
    await service.import_workbook(XLSX)
    favoritos_sin_resolver = clean_db.execute(
        "select count(*) as n from app.wishlist_item"
        " where is_favorite and card_id is null and source_option = 'galeria'"
    ).fetchone()["n"]
    assert favoritos_sin_resolver > 0


async def test_el_import_siembra_el_checklist_aunque_el_catalogo_este_caido(conn_factory, clean_db):
    """El dex number y el nombre vienen del Excel, no de TCGdex: la red
    caída no puede dejar el checklist vacío. Un ConnectTimeout en cualquier
    parte no debe abortar la corrida completa -- antes lo hacía, porque la
    excepción escapaba del resolver y `conn.commit()` nunca se alcanzaba."""
    service = ImportService(FakeCatalogPortSiempreFalla(), conn_factory)
    resumen = await service.import_workbook(XLSX)

    assert resumen.pokemon == 151
    assert len(repository.list_pokedex(clean_db)) == 151

    filas_1_y_2 = clean_db.execute(
        "select count(*) as n from app.wishlist_item"
        " where source_option in ('opcion_1', 'opcion_2')"
    ).fetchone()["n"]
    assert filas_1_y_2 == 0, "las opciones inalcanzables no deben guardarse"
    # 151 filas x (opción 1 + opción 2, ambas inalcanzables) como mínimo;
    # no se fija un número exacto porque también suma la galería.
    assert resumen.catalogo_inalcanzable >= 302


async def test_las_opciones_resueltas_conviven_con_las_inalcanzables(
    conn_factory, clean_db, tmp_path
):
    """Si el catálogo falla para una fila y responde para otra, la que
    respondió debe aterrizar normal y la que falló debe quedar ausente -- no
    guardada como si no hubiese resuelto."""
    path = _build_dos_pokemon_opcion1(tmp_path)
    port = FakeCatalogPortConFallas(cartas={"002": "sv03.5-002"}, fallan={"001"})
    service = ImportService(CatalogService(port, conn_factory), conn_factory)

    resumen = await service.import_workbook(path)

    filas = repository.list_wishlist(clean_db)
    assert len(filas) == 1, "solo debe aterrizar la opción que sí resolvió"
    assert filas[0]["dex_number"] == 2
    assert filas[0]["card_id"] == "sv03.5-002"
    assert resumen.catalogo_inalcanzable == 1
    assert resumen.sin_resolver == 0


async def test_el_no_match_sigue_distinto_del_inalcanzable(conn_factory, clean_db, tmp_path):
    """Las dos rutas se mantienen distintas: 'no existe' (el catálogo
    respondió que no) sigue guardando una fila sin resolver con su
    raw_text; 'no se pudo preguntar' no guarda nada."""
    path = _build_dos_pokemon_opcion1(tmp_path)
    # "001" falla (inalcanzable); "002" no está en `cartas` -> None real (no existe).
    port = FakeCatalogPortConFallas(fallan={"001"})
    service = ImportService(CatalogService(port, conn_factory), conn_factory)

    resumen = await service.import_workbook(path)

    filas = repository.list_wishlist(clean_db)
    assert len(filas) == 1, "el no-match sigue guardando su fila; el inalcanzable no"
    assert filas[0]["dex_number"] == 2
    assert filas[0]["card_id"] is None
    assert filas[0]["raw_text"] == "Clon 002/165"
    assert resumen.sin_resolver == 1
    assert resumen.catalogo_inalcanzable == 1


async def test_reimportar_despues_de_un_import_parcial_no_duplica(conn_factory, clean_db, tmp_path):
    """La garantía central de este round: si el catálogo estuvo caído para
    una opción, la corrida siguiente que sí resuelva no debe agregar una
    fila más encima -- el resultado final debe quedar igual que un import
    limpio de una sola pasada."""
    path = _build_dos_pokemon_opcion1(tmp_path)

    fallando = ImportService(
        CatalogService(FakeCatalogPortConFallas(fallan={"001", "002"}), conn_factory),
        conn_factory,
    )
    primero = await fallando.import_workbook(path)
    assert len(repository.list_wishlist(clean_db)) == 0
    assert primero.catalogo_inalcanzable == 2

    funcionando = ImportService(
        CatalogService(
            FakeCatalogPortConFallas(cartas={"001": "sv03.5-001", "002": "sv03.5-002"}),
            conn_factory,
        ),
        conn_factory,
    )
    segundo = await funcionando.import_workbook(path)

    filas = repository.list_wishlist(clean_db)
    assert len(filas) == 2, "debe quedar igual que un import limpio de una sola pasada"
    assert {f["card_id"] for f in filas} == {"sv03.5-001", "sv03.5-002"}
    assert segundo.catalogo_inalcanzable == 0
