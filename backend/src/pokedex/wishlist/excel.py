"""Lectura de `Pokedex_Viviente_151.xlsx` a filas estructuradas.

No sabe de cartas ni de sets: solo convierte celdas en datos. La resolución
contra el catálogo vive en `resolver.py`.

Layout de la hoja `Pokédex 151` (fila 3 son los encabezados, los datos
empiezan en la 4):
    A número de dex | B nombre | C ✔ (se ignora) | D opción elegida (se ignora)
    E opción 1 | F rareza | G valor USD
    I opción 2 | J rareza | K valor USD
    M opción 3 | N valor USD
    P opción 4
"""

from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from .models import ExcelOption, ExcelRow, GalleryRow

SHEET_DEX = "Pokédex 151"
SHEET_GALLERY = "Galería favoritos"

# (columna de la carta, columna del valor, nombre de la opción)
OPTION_COLUMNS = [
    ("E", "G", "opcion_1"),
    ("I", "K", "opcion_2"),
    ("M", "N", "opcion_3"),
    ("P", None, "opcion_4"),
]

VACIO = {"", "—", "-", "–", "None"}


def _text(cell) -> str:
    value = cell.value
    return "" if value is None else str(value).strip()


def _money(cell) -> Decimal | None:
    if cell is None:
        return None
    raw = _text(cell)
    if raw in VACIO:
        return None
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def parse_workbook(path: str | Path) -> tuple[list[ExcelRow], list[GalleryRow]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        rows = _parse_dex_sheet(workbook[SHEET_DEX])
        gallery = _parse_gallery_sheet(workbook[SHEET_GALLERY])
    finally:
        workbook.close()
    return rows, gallery


def _parse_dex_sheet(sheet) -> list[ExcelRow]:
    rows: list[ExcelRow] = []
    for excel_row in range(4, sheet.max_row + 1):
        numero = _text(sheet[f"A{excel_row}"])
        if not numero.isdigit():
            continue
        options = []
        for card_col, value_col, source_option in OPTION_COLUMNS:
            raw_text = _text(sheet[f"{card_col}{excel_row}"])
            if raw_text in VACIO:
                continue
            options.append(
                ExcelOption(
                    source_option=source_option,
                    raw_text=raw_text,
                    reference_value_usd=(
                        _money(sheet[f"{value_col}{excel_row}"]) if value_col else None
                    ),
                )
            )
        rows.append(
            ExcelRow(
                dex_number=int(numero),
                pokemon_name=_text(sheet[f"B{excel_row}"]),
                options=options,
            )
        )
    return rows


def _parse_gallery_sheet(sheet) -> list[GalleryRow]:
    gallery: list[GalleryRow] = []
    for excel_row in range(4, sheet.max_row + 1):
        numero = _text(sheet[f"A{excel_row}"])
        if not numero.isdigit():
            continue
        raw_text = _text(sheet[f"C{excel_row}"])
        if raw_text in VACIO:
            continue
        gallery.append(
            GalleryRow(
                dex_number=int(numero),
                pokemon_name=_text(sheet[f"B{excel_row}"]),
                raw_text=raw_text,
                reference_value_usd=_money(sheet[f"D{excel_row}"]),
            )
        )
    return gallery
